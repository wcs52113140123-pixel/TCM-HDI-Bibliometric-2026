"""
Day 25 Block A.5: Recover title/abstract + polish Excel form

Block A used cluster_assignments.parquet which lacks title/abstract.
This script:
  1. Auto-scans data/ + 02_data_integration/ + 03_topic_modeling/ for parquet with title+abstract+record_id
  2. Picks the file with highest record_id coverage
  3. Re-merges with the 50 sampled records
  4. Adds dropdown data validation (16 mechanism categories + unsure)
  5. Sets column widths, wraps title/abstract, freezes header row
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

REPO = Path(r"D:\Document\Research-Projects\TCM-HDI-Bibliometric-2026")
OUT_DIR = REPO / "results" / "supp_methods_s2"
form_path = OUT_DIR / "s2_annotator_form.xlsx"
master_path = OUT_DIR / "s2_master_keyed.xlsx"

# 16 Schema v3 categories (from Block A output) + 1 escape hatch
MECH_CATEGORIES = [
    'CYP_inhibition', 'CYP_induction',
    'UGT_inhibition', 'UGT_induction',
    'P-gp_inhibition', 'P-gp_induction',
    'transporter_modulation',
    'absorption_alteration',
    'protein_binding_displacement',
    'receptor_synergism', 'receptor_antagonism',
    'synergistic_efficacy', 'antagonistic_efficacy',
    'additive_toxicity',
    'organ_toxicity_modulation',
    'signaling_pathway_modulation',
    'unsure_ambiguous',
]

# ---- Step 1: Load current S2 ----
form_df = pd.read_excel(form_path)
master_df = pd.read_excel(master_path)
record_ids = set(form_df['record_id'].tolist())
print(f"Need title/abstract for {len(record_ids)} record_ids")

# ---- Step 2: Discover parquet with title+abstract+record_id ----
search_dirs = [REPO / "data", REPO / "02_data_integration", REPO / "03_topic_modeling", REPO]
candidates = []
seen = set()
for sd in search_dirs:
    if not sd.exists(): continue
    for pq in sd.rglob("*.parquet"):
        if pq in seen: continue
        seen.add(pq)
        try:
            cols = pd.read_parquet(pq, columns=None).columns.tolist()
        except Exception:
            continue
        has_title = any('title' in c.lower() for c in cols)
        has_abs   = any('abstract' in c.lower() for c in cols)
        has_recid = any(c.lower() in ['record_id', 'paper_id'] for c in cols)
        if has_title and has_abs and has_recid:
            size_mb = pq.stat().st_size / 1024 / 1024
            candidates.append((pq, cols, size_mb))

if not candidates:
    print("[ERROR] No parquet found with title+abstract+record_id columns.")
    print("Please tell me corpus master path manually.")
    raise SystemExit(1)

def coverage(pq):
    for col_name in ['record_id', 'paper_id']:
        try:
            df_sub = pd.read_parquet(pq, columns=[col_name])
            return int(df_sub[col_name].isin(record_ids).sum())
        except Exception:
            continue
    return 0

print(f"\nFound {len(candidates)} candidate file(s):")
for pq, _, sz in candidates:
    cov = coverage(pq)
    print(f"  {pq.relative_to(REPO)} ({sz:.1f} MB, coverage {cov}/{len(record_ids)})")

candidates.sort(key=lambda x: -coverage(x[0]))
chosen, _, _ = candidates[0]
print(f"\nUsing: {chosen.relative_to(REPO)}")

corpus = pd.read_parquet(chosen)
join_col = next((c for c in corpus.columns if c.lower() in ['record_id', 'paper_id']), None)
title_col = next((c for c in corpus.columns if c.lower() == 'title'), None)
abs_col = next((c for c in corpus.columns if c.lower() == 'abstract'), None)
year_col = next((c for c in corpus.columns if c.lower() in ['year', 'publication_year']), None)
print(f"  Cols: join={join_col}, title={title_col}, abstract={abs_col}, year={year_col}")

cols_to_get = [c for c in [join_col, title_col, abs_col, year_col] if c]
corpus_sub = corpus[corpus[join_col].isin(record_ids)][cols_to_get].drop_duplicates(subset=[join_col])
print(f"  Matched: {len(corpus_sub)}/{len(record_ids)} records")

# ---- Step 3: Re-merge ----
form_df = form_df.drop(columns=['year'], errors='ignore')
form_df = form_df.merge(corpus_sub, left_on='record_id', right_on=join_col, how='left')
master_df = master_df.drop(columns=['year'], errors='ignore')
master_df = master_df.merge(corpus_sub, left_on='record_id', right_on=join_col, how='left')

def reorder(df, include_llm=False):
    cols = ['annotator_record_id', 'record_id']
    if year_col: cols.append(year_col)
    if title_col: cols.append(title_col)
    if abs_col: cols.append(abs_col)
    cols += ['your_mechanism_label', 'your_alternative_label_optional',
             'your_confidence_low_med_high', 'your_notes_optional']
    if include_llm:
        cols.append('llm_mechanism_label')
        if 'llm_confidence' in df.columns:
            cols.append('llm_confidence')
    return df[[c for c in cols if c in df.columns]]

form_df = reorder(form_df, include_llm=False)
master_df = reorder(master_df, include_llm=True)
form_df.to_excel(form_path, index=False)
master_df.to_excel(master_path, index=False)
print(f"\n[+] Re-merged + saved {form_path.name}")
print(f"[+] Re-merged + saved {master_path.name}")

# ---- Step 4: Polish annotator form (dropdowns + widths + freeze) ----
wb = load_workbook(form_path)
ws = wb.active

# Hidden sheet for category list (list-string exceeds 255 char limit)
if "_categories" in wb.sheetnames:
    del wb["_categories"]
ws_cat = wb.create_sheet(title="_categories", index=1)
ws_cat['A1'] = "MECHANISM_CATEGORIES"
ws_cat['A1'].font = Font(bold=True)
for i, cat in enumerate(MECH_CATEGORIES):
    ws_cat.cell(row=i+2, column=1, value=cat)
ws_cat['C1'] = "CONFIDENCE_LEVELS"
ws_cat['C1'].font = Font(bold=True)
for i, lvl in enumerate(['low', 'medium', 'high']):
    ws_cat.cell(row=i+2, column=3, value=lvl)
ws_cat.sheet_state = 'hidden'

header_row = [cell.value for cell in ws[1]]
def col_letter(name):
    return get_column_letter(header_row.index(name) + 1) if name in header_row else None

# Add dropdowns
mech_ref = f"_categories!$A$2:$A${len(MECH_CATEGORIES)+1}"
conf_ref = "_categories!$C$2:$C$4"

for col_name in ['your_mechanism_label', 'your_alternative_label_optional']:
    if col_letter(col_name):
        dv = DataValidation(type="list", formula1=f"={mech_ref}", allow_blank=True)
        dv.add(f"{col_letter(col_name)}2:{col_letter(col_name)}51")
        ws.add_data_validation(dv)

if col_letter('your_confidence_low_med_high'):
    dv = DataValidation(type="list", formula1=f"={conf_ref}", allow_blank=True)
    dv.add(f"{col_letter('your_confidence_low_med_high')}2:{col_letter('your_confidence_low_med_high')}51")
    ws.add_data_validation(dv)

# Column widths
widths = {
    'annotator_record_id': 12, 'record_id': 14, 'year': 8,
    'title': 50, 'abstract': 90,
    'your_mechanism_label': 28, 'your_alternative_label_optional': 28,
    'your_confidence_low_med_high': 18, 'your_notes_optional': 30,
}
for name, w in widths.items():
    letter = col_letter(name)
    if letter:
        ws.column_dimensions[letter].width = w

# Wrap text for title + abstract
for row in ws.iter_rows(min_row=2, max_row=51):
    for cell in row:
        col_name = header_row[cell.column - 1]
        if col_name in ['title', 'abstract']:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

# Header style
fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 35

# Row height for content (so abstract is visible)
for row_num in range(2, 52):
    ws.row_dimensions[row_num].height = 100

wb.save(form_path)
print(f"[+] {form_path.name}: dropdowns + widths + frozen header applied")

# ---- Sanity check ----
print(f"\n=== Sanity check: first 2 records ===")
for i, row in form_df.head(2).iterrows():
    print(f"\n  [{row['annotator_record_id']}] record_id={row['record_id']}")
    if year_col: print(f"    Year: {row.get(year_col)}")
    if title_col: print(f"    Title: {str(row[title_col])[:120]}")
    if abs_col:
        ab = str(row[abs_col])
        print(f"    Abstract: {ab[:250]}{'...' if len(ab)>250 else ''}")

print(f"\nDone. Send {form_path.name} to shimei. Keep {master_path.name} private for Day 28 kappa.")